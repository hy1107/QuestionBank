import pytest
import tempfile
import os
from parser.excel_handler import questions_to_excel, excel_to_questions

SAMPLE_QUESTIONS = [
    {
        'original_no': '1',
        'question': '測試題目一',
        'option_a': '選A',
        'option_b': '選B',
        'option_c': '選C',
        'option_d': '選D',
        'answer': 'A',
        'subject': '',
        'chapter': '',
        'difficulty': 'medium',
        'source_file': 'test.pdf',
    },
    {
        'original_no': '2',
        'question': '測試題目二',
        'option_a': '甲',
        'option_b': '乙',
        'option_c': '丙',
        'option_d': '丁',
        'answer': 'C',
        'subject': '',
        'chapter': '',
        'difficulty': 'easy',
        'source_file': 'test.pdf',
    },
]

def test_questions_to_excel_creates_file(tmp_path):
    output = str(tmp_path / 'output.xlsx')
    questions_to_excel(SAMPLE_QUESTIONS, output)
    assert os.path.exists(output)

def test_questions_to_excel_has_correct_headers(tmp_path):
    output = str(tmp_path / 'output.xlsx')
    questions_to_excel(SAMPLE_QUESTIONS, output)
    from openpyxl import load_workbook
    wb = load_workbook(output)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, 8)]
    assert headers == ['題號', '題目', '選項A', '選項B', '選項C', '選項D', '答案']

def test_questions_to_excel_has_correct_data(tmp_path):
    output = str(tmp_path / 'output.xlsx')
    questions_to_excel(SAMPLE_QUESTIONS, output)
    from openpyxl import load_workbook
    wb = load_workbook(output)
    ws = wb.active
    assert ws.cell(2, 1).value == '1'
    assert ws.cell(2, 2).value == '測試題目一'
    assert ws.cell(2, 7).value == 'A'

def test_excel_to_questions_roundtrip(tmp_path):
    output = str(tmp_path / 'roundtrip.xlsx')
    questions_to_excel(SAMPLE_QUESTIONS, output)
    loaded = excel_to_questions(output)
    assert len(loaded) == 2
    assert loaded[0]['original_no'] == '1'
    assert loaded[0]['question'] == '測試題目一'
    assert loaded[0]['answer'] == 'A'
    assert loaded[1]['option_c'] == '丙'

def test_excel_to_questions_validates_answer_column(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(['題號', '題目', '選項A', '選項B', '選項C', '選項D', '答案'])
    ws.append(['1', '題目', 'A', 'B', 'C', 'D', 'X'])  # invalid answer
    bad_file = str(tmp_path / 'bad.xlsx')
    wb.save(bad_file)
    with pytest.raises(ValueError, match='答案欄位'):
        excel_to_questions(bad_file)
