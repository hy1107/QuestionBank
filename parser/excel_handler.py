from openpyxl import Workbook, load_workbook

HEADERS = ['題號', '題目', '選項A', '選項B', '選項C', '選項D', '答案']

def questions_to_excel(questions, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = '題庫'
    ws.append(HEADERS)
    for q in questions:
        ws.append([
            q['original_no'],
            q['question'],
            q['option_a'],
            q['option_b'],
            q['option_c'],
            q['option_d'],
            q['answer'],
        ])
    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    wb.save(output_path)

def excel_to_questions(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # Skip header row
    questions = []
    for i, row in enumerate(rows[1:], start=2):
        if not row or not any(row):
            continue
        no, question, a, b, c, d, answer = (str(v).strip() if v is not None else '' for v in row[:7])
        if answer.upper() not in ('A', 'B', 'C', 'D'):
            raise ValueError(f'答案欄位第 {i} 列格式錯誤："{answer}"，必須為 A/B/C/D')
        questions.append({
            'original_no': no,
            'question': question,
            'option_a': a,
            'option_b': b,
            'option_c': c,
            'option_d': d,
            'answer': answer.upper(),
            'subject': '',
            'chapter': '',
            'difficulty': 'medium',
            'source_file': filepath,
        })
    return questions
